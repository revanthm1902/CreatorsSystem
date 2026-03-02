"""
AryVerse Creators System — Supabase Data Exporter
Generates: CreatorsSystem_Export.xlsx  (multi-sheet workbook)
Requires:  pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ─────────────────────────────────
# DATA  (fetched from Supabase MCP)
# ─────────────────────────────────

PROFILES = [
    {"employee_id":"AV2026001","full_name":"Aryan Goyal","role":"Director","department":"Founder","email":None,"phone":None,"date_of_birth":None,"total_tokens":0,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":None,"github_url":None,"resume_url":None,"is_temporary_password":False,"created_at":"2026-02-26 16:17","updated_at":"2026-03-02 07:20"},
    {"employee_id":"AV2026002","full_name":"Revanth Modalavalasa","role":"Admin","department":"Admin","email":"revanthm051@gmail.com","phone":"7337426334","date_of_birth":"2006-02-19","total_tokens":1001,"banked_minutes":0,"token_credit_balance":0.1,"linkedin_url":"https://revanthm.vercel.app/","github_url":"https://github.com/revanthm1902","resume_url":"https://drive.google.com/file/d/1AzZ8LE9B6So3RJFd4YtSxmZcxtjTfiPp/view","is_temporary_password":False,"created_at":"2026-02-26 17:09","updated_at":"2026-03-02 13:25"},
    {"employee_id":"AV2026004","full_name":"Frank Mathew Sajan","role":"Admin","department":"Admin","email":"frank@aryverse.com","phone":"+91 9497182886","date_of_birth":"2005-06-01","total_tokens":1000,"banked_minutes":0,"token_credit_balance":0.6,"linkedin_url":"https://www.linkedin.com/in/frankmathewsajan/","github_url":"https://github.com/frankmathewsajan","resume_url":"https://itsfrank.web.app/resume","is_temporary_password":False,"created_at":"2026-02-27 17:39","updated_at":"2026-03-02 08:45"},
    {"employee_id":"AV2026005","full_name":"Ganesh Sai Midhun Gullapalli","role":"User","department":"AI Engineer","email":"midhungullapalli9@gmail.com","phone":"+91 9110581559","date_of_birth":"2006-01-17","total_tokens":6,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":"https://www.linkedin.com/in/midhun-gullapalli/","github_url":"https://github.com/Midhun-gg","resume_url":"https://docs.google.com/document/d/12gqQGSXsY_iDD_F9X0EGHcLmW05ICEJCnyZFQAGgrxY","is_temporary_password":False,"created_at":"2026-02-28 11:45","updated_at":"2026-03-02 16:09"},
    {"employee_id":"AV2026006","full_name":"Datta Sai Mithun Gullapalli","role":"User","department":"AI Engineer","email":"mithung.2077@gmail.com","phone":"9492841569","date_of_birth":"2006-01-17","total_tokens":6,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":"https://www.linkedin.com/in/mithun-gullapalli-003787290/","github_url":"https://github.com/Mithun20222","resume_url":"https://drive.google.com/file/d/16_vYeilAGL39YzE3vduXK59YvORjm7xP/view","is_temporary_password":False,"created_at":"2026-02-28 11:54","updated_at":"2026-03-02 16:10"},
    {"employee_id":"AV2026016","full_name":"Dhanish","role":"User","department":"AI Engineer","email":"dhanishraja2007@gmail.com","phone":"9398497278","date_of_birth":None,"total_tokens":0,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":"https://www.linkedin.com/in/raja-dhanish-475777293","github_url":"https://github.com/dhanish-raja","resume_url":"https://1drv.ms/b/c/568d4a436c25e46e/IQDtzq015xioQobdJDle5UoMASP2_FycV4N0YGNijpae7Z8","is_temporary_password":False,"created_at":"2026-02-28 20:37","updated_at":"2026-03-02 17:08"},
    {"employee_id":"AV2026017","full_name":"Jyothi Reddy.Pula","role":"User","department":"Web Dev","email":"pulajyothireddy09@gmail.com","phone":"7013565588","date_of_birth":"2006-08-09","total_tokens":6,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":"https://www.linkedin.com/in/jyothireddy-pula-5b3a01337/","github_url":"https://github.com/Jyothireddy-pula","resume_url":"https://drive.google.com/file/d/1hkiKo1tmtu-ZvUQ2QpSzxvbOxjV8_z-K/view","is_temporary_password":False,"created_at":"2026-02-28 20:44","updated_at":"2026-03-02 16:24"},
    {"employee_id":"AV2026018","full_name":"Meghana","role":"User","department":"Non-Technical","email":None,"phone":None,"date_of_birth":None,"total_tokens":100,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":None,"github_url":None,"resume_url":None,"is_temporary_password":False,"created_at":"2026-02-28 20:45","updated_at":"2026-03-02 09:07"},
    {"employee_id":"AV2026019","full_name":"Aurang","role":"User","department":"Non-Technical","email":None,"phone":None,"date_of_birth":None,"total_tokens":0,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":None,"github_url":None,"resume_url":None,"is_temporary_password":False,"created_at":"2026-02-28 20:46","updated_at":"2026-03-02 09:07"},
    {"employee_id":"AV2026020","full_name":"Kavya","role":"User","department":"Management Executive","email":"kavyamallidi2324@gmail.com","phone":"9153733999","date_of_birth":"2006-04-04","total_tokens":0,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":None,"github_url":None,"resume_url":None,"is_temporary_password":False,"created_at":"2026-02-28 20:48","updated_at":"2026-03-02 09:07"},
    {"employee_id":"AV2026021","full_name":"Kishor","role":"User","department":"Non-Technical","email":"kishor09.official@gmail.com","phone":"9360099023","date_of_birth":"2007-08-09","total_tokens":0,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":None,"github_url":None,"resume_url":None,"is_temporary_password":False,"created_at":"2026-02-28 20:49","updated_at":"2026-03-02 09:07"},
    {"employee_id":"AV2026022","full_name":"Joseph","role":"User","department":"Web Dev","email":None,"phone":None,"date_of_birth":None,"total_tokens":0,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":None,"github_url":None,"resume_url":None,"is_temporary_password":False,"created_at":"2026-02-28 20:49","updated_at":"2026-03-02 09:07"},
    {"employee_id":"AV2026023","full_name":"Himanvi Damarla","role":"User","department":"Management Executive","email":"damarlahimanvi@gmail.com","phone":"6309664715","date_of_birth":"2005-04-16","total_tokens":0,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":"https://www.linkedin.com/in/himanvi-damarla/","github_url":"https://github.com/himanvidamarla","resume_url":"https://drive.google.com/file/d/1Iw8ZxtSZhuqsvf5ES_7dltTvrSeoeyvJ/view","is_temporary_password":False,"created_at":"2026-02-28 20:52","updated_at":"2026-03-02 16:08"},
    {"employee_id":"AV2026024","full_name":"Teja","role":"User","department":"Web Dev","email":None,"phone":None,"date_of_birth":None,"total_tokens":3,"banked_minutes":0,"token_credit_balance":0,"linkedin_url":None,"github_url":None,"resume_url":None,"is_temporary_password":False,"created_at":"2026-02-28 20:53","updated_at":"2026-03-02 09:07"},
]

TASKS = [
    {"title":"Apply for copyright","created_by":"Aryan Goyal","assigned_to":"Himanvi Damarla","assignee_id":"AV2026023","status":"Pending","tokens":5,"director_approved":True,"deadline":"2026-03-05 18:29","original_deadline":None,"submitted_at":None,"approved_at":None,"submission_note":None,"admin_feedback":None,"pow_url":None,"issue_state":None,"created_at":"2026-03-01 16:16"},
    {"title":"UI of WebApp for AryVerse","created_by":"Revanth Modalavalasa","assigned_to":"Kavya","assignee_id":"AV2026020","status":"Pending","tokens":5,"director_approved":True,"deadline":"2026-03-04 18:29","original_deadline":None,"submitted_at":None,"approved_at":None,"submission_note":None,"admin_feedback":None,"pow_url":None,"issue_state":None,"created_at":"2026-03-01 14:25"},
    {"title":"Add sound to the video","created_by":"Aryan Goyal","assigned_to":"Kishor","assignee_id":"AV2026021","status":"Pending","tokens":5,"director_approved":True,"deadline":"2026-03-07 18:29","original_deadline":None,"submitted_at":"2026-03-02 12:16","approved_at":None,"submission_note":None,"admin_feedback":None,"pow_url":None,"issue_state":None,"created_at":"2026-03-01 12:57"},
    {"title":"Spatial Extraction","created_by":"Frank Mathew Sajan","assigned_to":"Ganesh Sai Midhun Gullapalli","assignee_id":"AV2026005","status":"Completed","tokens":6,"director_approved":True,"deadline":"2026-03-03 18:29","original_deadline":None,"submitted_at":"2026-03-02 08:05","approved_at":"2026-03-02 08:12","submission_note":"Evaluated multiple monocular human pose estimation frameworks and selected the most suitable option. Documented progress by updating the corresponding GitHub issue.","admin_feedback":"Great job!","pow_url":"https://github.com/AryVerse/ai-core/issues/1","issue_state":"closed","created_at":"2026-03-01 10:25"},
    {"title":"Linguistic Translation","created_by":"Frank Mathew Sajan","assigned_to":"Datta Sai Mithun Gullapalli","assignee_id":"AV2026006","status":"Completed","tokens":6,"director_approved":True,"deadline":"2026-03-03 18:29","original_deadline":None,"submitted_at":"2026-03-02 08:05","approved_at":"2026-03-02 08:45","submission_note":"Investigated methodologies for utilizing LLMs to output specific numerical shape adjustments for a standardized human baseline.","admin_feedback":"We will move forward with Method 5 (Hybrid LLM). Closing this research ticket, opening a new issue for prototyping.","pow_url":"https://github.com/AryVerse/ai-core/issues/2","issue_state":"closed","created_at":"2026-03-01 10:25"},
    {"title":"Basic Mesh Creation","created_by":"Revanth Modalavalasa","assigned_to":"Meghana","assignee_id":"AV2026018","status":"Pending","tokens":5,"director_approved":True,"deadline":"2026-03-03 18:29","original_deadline":None,"submitted_at":None,"approved_at":None,"submission_note":None,"admin_feedback":None,"pow_url":None,"issue_state":None,"created_at":"2026-03-01 09:29"},
    {"title":"Customer Support Chatbot Integration","created_by":"Revanth Modalavalasa","assigned_to":"Joseph","assignee_id":"AV2026022","status":"Pending","tokens":10,"director_approved":True,"deadline":"2026-03-04 18:29","original_deadline":None,"submitted_at":None,"approved_at":None,"submission_note":None,"admin_feedback":None,"pow_url":"https://github.com/AryVerse/web-platform/issues/1","issue_state":"open","created_at":"2026-03-01 07:37"},
    {"title":"Social Media & Store Link Migration","created_by":"Revanth Modalavalasa","assigned_to":"Teja","assignee_id":"AV2026024","status":"Completed","tokens":2,"director_approved":True,"deadline":"2026-03-02 06:30","original_deadline":None,"submitted_at":"2026-03-02 04:53","approved_at":"2026-03-02 04:58","submission_note":"Added social media icons in footer. Created a Coming Soon page. Improved styling of Privacy Policy, Terms of Service, and Cookie Policy pages.","admin_feedback":"Verified. Good work.","pow_url":"https://github.com/AryVerse/web-platform/issues/2","issue_state":"closed","created_at":"2026-03-01 07:34"},
    {"title":"Navbar Expansion & Content Architecture","created_by":"Revanth Modalavalasa","assigned_to":"Jyothi Reddy.Pula","assignee_id":"AV2026017","status":"Completed","tokens":6,"director_approved":True,"deadline":"2026-03-03 18:29","original_deadline":None,"submitted_at":"2026-03-02 09:46","approved_at":"2026-03-02 09:48","submission_note":"Implemented expanded navbar layout, updated content structure for main sections, verified responsive behavior and basic navigation links.","admin_feedback":"Good work.","pow_url":"https://github.com/AryVerse/web-platform/issues/3","issue_state":"closed","created_at":"2026-03-01 07:33"},
    {"title":"Humanoid Mesh Adjusting AI","created_by":"Frank Mathew Sajan","assigned_to":"Dhanish","assignee_id":"AV2026016","status":"Pending","tokens":20,"director_approved":True,"deadline":"2026-03-08 18:29","original_deadline":None,"submitted_at":None,"approved_at":None,"submission_note":None,"admin_feedback":None,"pow_url":None,"issue_state":None,"created_at":"2026-03-01 05:27"},
]

POINTS_LOG = [
    {"employee_id":"AV2026017","full_name":"Jyothi Reddy.Pula","task_title":"Navbar Expansion & Content Architecture","tokens_awarded":6,"reason":"Task completed on time","created_at":"2026-03-02 09:48"},
    {"employee_id":"AV2026006","full_name":"Datta Sai Mithun Gullapalli","task_title":"Linguistic Translation","tokens_awarded":6,"reason":"Task completed on time","created_at":"2026-03-02 08:45"},
    {"employee_id":"AV2026005","full_name":"Ganesh Sai Midhun Gullapalli","task_title":"Spatial Extraction","tokens_awarded":6,"reason":"Task completed on time","created_at":"2026-03-02 08:12"},
    {"employee_id":"AV2026024","full_name":"Teja","task_title":"Social Media & Store Link Migration","tokens_awarded":3,"reason":"Task completed on time (+1 bonus)","created_at":"2026-03-02 04:58"},
]

ACTIVITY_LOG = [
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Joseph","task_title":"Customer Support Chatbot Integration","action_type":"task_reassigned","message":"Revanth Modalavalasa reassigned task \"Customer Support Chatbot Integration\" back to Joseph for rework","created_at":"2026-03-02 16:40"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Joseph","task_title":"Customer Support Chatbot Integration","action_type":"director_approved_task","message":"Aryan Goyal approved task \"Customer Support Chatbot Integration\" for users","created_at":"2026-03-02 14:05"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Joseph","task_title":"Customer Support Chatbot Integration","action_type":"task_assigned","message":"Revanth Modalavalasa edited task \"Customer Support Chatbot Integration\" (pending director re-approval)","created_at":"2026-03-02 12:30"},
    {"actor_name":"Kishor","actor_role":"User","target_user":"Kishor","task_title":"Add sound to the video","action_type":"task_marked_done","message":"Kishor submitted task \"Add sound to the video\" for review","created_at":"2026-03-02 12:16"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Himanvi Damarla","task_title":"Apply for copyright","action_type":"task_assigned","message":"Aryan Goyal edited task \"Apply for copyright\"","created_at":"2026-03-02 11:38"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Jyothi Reddy.Pula","task_title":"Navbar Expansion & Content Architecture","action_type":"task_approved","message":"Revanth Modalavalasa approved task \"Navbar Expansion & Content Architecture\" for Jyothi Reddy.Pula (+6 tokens)","created_at":"2026-03-02 09:48"},
    {"actor_name":"Jyothi Reddy.Pula","actor_role":"User","target_user":"Jyothi Reddy.Pula","task_title":"Navbar Expansion & Content Architecture","action_type":"task_marked_done","message":"Jyothi Reddy.Pula submitted task \"Navbar Expansion & Content Architecture\" for review","created_at":"2026-03-02 09:46"},
    {"actor_name":"Frank Mathew Sajan","actor_role":"Admin","target_user":"Datta Sai Mithun Gullapalli","task_title":"Linguistic Translation","action_type":"task_approved","message":"Frank Mathew Sajan approved task \"Linguistic Translation\" for Mithun (+6 tokens)","created_at":"2026-03-02 08:45"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Ganesh Sai Midhun Gullapalli","task_title":"Spatial Extraction","action_type":"task_approved","message":"Revanth Modalavalasa approved task \"Spatial Extraction\" for Midhun (+6 tokens)","created_at":"2026-03-02 08:12"},
    {"actor_name":"Ganesh Sai Midhun Gullapalli","actor_role":"User","target_user":"Ganesh Sai Midhun Gullapalli","task_title":"Spatial Extraction","action_type":"task_marked_done","message":"Midhun submitted task \"Spatial Extraction\" for review","created_at":"2026-03-02 08:05"},
    {"actor_name":"Datta Sai Mithun Gullapalli","actor_role":"User","target_user":"Datta Sai Mithun Gullapalli","task_title":"Linguistic Translation","action_type":"task_marked_done","message":"Mithun submitted task \"Linguistic Translation\" for review","created_at":"2026-03-02 08:05"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Teja","task_title":"Social Media & Store Link Migration","action_type":"task_approved","message":"Revanth Modalavalasa approved task \"Social Media & Store Link Migration\" for Teja (+2 tokens +1 bonus)","created_at":"2026-03-02 04:58"},
    {"actor_name":"Teja","actor_role":"User","target_user":"Teja","task_title":"Social Media & Store Link Migration","action_type":"task_marked_done","message":"Teja submitted task \"Social Media & Store Link Migration\" for review","created_at":"2026-03-02 04:53"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Jyothi Reddy.Pula","task_title":"Navbar Expansion & Content Architecture","action_type":"task_reassigned","message":"Revanth Modalavalasa reassigned task \"Navbar Expansion & Content Architecture\" back to Jyothi Reddy.Pula for rework","created_at":"2026-03-01 17:01"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Himanvi Damarla","task_title":"Apply for copyright","action_type":"task_assigned","message":"Aryan Goyal assigned task \"Apply for copyright\" to Himanvi Damarla","created_at":"2026-03-01 16:16"},
    {"actor_name":"Jyothi Reddy.Pula","actor_role":"User","target_user":"Jyothi Reddy.Pula","task_title":"Navbar Expansion & Content Architecture","action_type":"task_marked_done","message":"Jyothi Reddy.Pula submitted task \"Navbar Expansion & Content Architecture\" for review","created_at":"2026-03-01 16:14"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":None,"task_title":None,"action_type":"custom_message","message":"Once you are done with tasks, make sure to \"Mark as done\", as you'll be eligible to be rewarded for early completion","created_at":"2026-03-01 14:41"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Kavya","task_title":"UI of WebApp for AryVerse","action_type":"director_approved_task","message":"Aryan Goyal approved task \"UI of WebApp for AryVerse\" for users","created_at":"2026-03-01 14:38"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Kavya","task_title":"UI of WebApp for AryVerse","action_type":"task_assigned","message":"Revanth Modalavalasa assigned task \"UI of WebApp for AryVerse\" to Kavya","created_at":"2026-03-01 14:25"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Kishor","task_title":"Add sound to the video","action_type":"task_assigned","message":"Aryan Goyal assigned task \"Add sound to the video\" to Kishor","created_at":"2026-03-01 12:57"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Datta Sai Mithun Gullapalli","task_title":"Linguistic Translation","action_type":"director_approved_task","message":"Aryan Goyal approved task \"Linguistic Translation\" for users","created_at":"2026-03-01 12:30"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Ganesh Sai Midhun Gullapalli","task_title":"Spatial Extraction","action_type":"director_approved_task","message":"Aryan Goyal approved task \"Spatial Extraction\" for users","created_at":"2026-03-01 12:30"},
    {"actor_name":"Frank Mathew Sajan","actor_role":"Admin","target_user":"Ganesh Sai Midhun Gullapalli","task_title":"Spatial Extraction","action_type":"task_assigned","message":"Frank Mathew Sajan assigned task \"Spatial Extraction\" to Midhun","created_at":"2026-03-01 10:25"},
    {"actor_name":"Frank Mathew Sajan","actor_role":"Admin","target_user":"Datta Sai Mithun Gullapalli","task_title":"Linguistic Translation","action_type":"task_assigned","message":"Frank Mathew Sajan assigned task \"Linguistic Translation\" to Mithun","created_at":"2026-03-01 10:25"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Meghana","task_title":"Basic Mesh Creation","action_type":"director_approved_task","message":"Aryan Goyal approved task \"Basic Mesh Creation\" for users","created_at":"2026-03-01 09:29"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Meghana","task_title":"Basic Mesh Creation","action_type":"task_assigned","message":"Revanth Modalavalasa assigned task \"Basic Mesh Creation\" to Meghana","created_at":"2026-03-01 09:29"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Jyothi Reddy.Pula","task_title":"Navbar Expansion & Content Architecture","action_type":"director_approved_task","message":"Aryan Goyal approved task \"Navbar Expansion & Content Architecture\" for users","created_at":"2026-03-01 08:10"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Teja","task_title":"Social Media & Store Link Migration","action_type":"director_approved_task","message":"Aryan Goyal approved task \"Social Media & Store Link Migration\" for users","created_at":"2026-03-01 08:10"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Joseph","task_title":"Customer Support Chatbot Integration","action_type":"director_approved_task","message":"Aryan Goyal approved task \"Customer Support Chatbot Integration\" for users","created_at":"2026-03-01 08:10"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Joseph","task_title":"Customer Support Chatbot Integration","action_type":"task_assigned","message":"Revanth Modalavalasa assigned task \"Customer Support Chatbot Integration\" to Joseph","created_at":"2026-03-01 07:37"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Teja","task_title":"Social Media & Store Link Migration","action_type":"task_assigned","message":"Revanth Modalavalasa assigned task \"Social Media & Store Link Migration\" to Teja","created_at":"2026-03-01 07:34"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Jyothi Reddy.Pula","task_title":"Navbar Expansion & Content Architecture","action_type":"task_assigned","message":"Revanth Modalavalasa assigned task \"Navbar Expansion & Content Architecture\" to Jyothi Reddy","created_at":"2026-03-01 07:33"},
    {"actor_name":"Aryan Goyal","actor_role":"Director","target_user":"Dhanish","task_title":"Humanoid Mesh Adjusting AI","action_type":"director_approved_task","message":"Aryan Goyal approved task \"Humanoid Mesh Adjusting AI\" for users","created_at":"2026-03-01 05:41"},
    {"actor_name":"Frank Mathew Sajan","actor_role":"Admin","target_user":"Dhanish","task_title":"Humanoid Mesh Adjusting AI","action_type":"task_assigned","message":"Frank Mathew Sajan assigned task \"Humanoid Mesh Adjusting AI\" to Dhanish","created_at":"2026-03-01 05:27"},
    {"actor_name":"Frank Mathew Sajan","actor_role":"Admin","target_user":None,"task_title":None,"action_type":"task_deleted","message":"Frank Mathew Sajan deleted task \"Parametric Engine\"","created_at":"2026-03-01 05:26"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Teja","task_title":None,"action_type":"user_added","message":"Revanth Modalavalasa added Teja as User","created_at":"2026-02-28 20:53"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Himanvi Damarla","task_title":None,"action_type":"user_added","message":"Revanth Modalavalasa added Himanvi as User","created_at":"2026-02-28 20:52"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Joseph","task_title":None,"action_type":"user_added","message":"Revanth Modalavalasa added Joseph as User","created_at":"2026-02-28 20:49"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Kishor","task_title":None,"action_type":"user_added","message":"Revanth Modalavalasa added Kishor as User","created_at":"2026-02-28 20:49"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Aurang","task_title":None,"action_type":"user_added","message":"Revanth Modalavalasa added Aurang as User","created_at":"2026-02-28 20:46"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Dhanish","task_title":None,"action_type":"user_added","message":"Revanth Modalavalasa added Dhanish as User","created_at":"2026-02-28 20:37"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":"Kavya","task_title":None,"action_type":"user_added","message":"Revanth Modalavalasa added Kavya as User","created_at":"2026-02-28 16:57"},
    {"actor_name":"Revanth Modalavalasa","actor_role":"Admin","target_user":None,"task_title":None,"action_type":"custom_message","message":"Done with testing! Time to go Live.","created_at":"2026-02-27 18:13"},
]

DEPARTMENT_ACCESS = [
    {"user_emp_id":"AV2026023","user_name":"Himanvi Damarla","department":None,"can_view_department":"AI Engineer","granted_by_name":"Revanth Modalavalasa","created_at":"2026-03-02 11:06"},
    {"user_emp_id":"AV2026023","user_name":"Himanvi Damarla","department":None,"can_view_department":"App Dev","granted_by_name":"Revanth Modalavalasa","created_at":"2026-03-02 11:06"},
    {"user_emp_id":"AV2026023","user_name":"Himanvi Damarla","department":None,"can_view_department":"Management Executive","granted_by_name":"Revanth Modalavalasa","created_at":"2026-03-02 11:06"},
    {"user_emp_id":"AV2026023","user_name":"Himanvi Damarla","department":None,"can_view_department":"Non-Technical","granted_by_name":"Revanth Modalavalasa","created_at":"2026-03-02 11:06"},
    {"user_emp_id":"AV2026020","user_name":"Kavya","department":None,"can_view_department":"Web Dev","granted_by_name":"Revanth Modalavalasa","created_at":"2026-03-02 11:06"},
    {"user_emp_id":"AV2026023","user_name":"Himanvi Damarla","department":None,"can_view_department":"Web Dev","granted_by_name":"Revanth Modalavalasa","created_at":"2026-03-02 11:06"},
]

PASSWORD_RESET_REQUESTS = [
    {"email":"ljremi@gmail.com","status":"pending","resolved_by_name":None,"created_at":"2026-03-02 09:56","resolved_at":None},
    {"email":"dhanishraja2007@gmail.com","status":"dismissed","resolved_by_name":None,"created_at":"2026-02-28 20:35","resolved_at":"2026-03-01 04:42"},
]


# ─────────────────────────────────
# STYLING HELPERS
# ─────────────────────────────────

# Color palette
PURPLE     = "7C3AED"   # header bg
PURPLE_LT  = "EDE9FE"   # alt row
WHITE      = "FFFFFF"
DARK       = "1E1B4B"
GREEN_BG   = "D1FAE5"
YELLOW_BG  = "FEF9C3"
RED_BG     = "FEE2E2"
BLUE_BG    = "DBEAFE"
GRAY_BG    = "F3F4F6"

STATUS_COLORS = {
    "Completed": GREEN_BG,
    "Pending":   YELLOW_BG,
    "Rejected":  RED_BG,
    "Under Review": BLUE_BG,
    "pending":   YELLOW_BG,
    "approved":  GREEN_BG,
    "dismissed": GRAY_BG,
}

ROLE_COLORS = {
    "Director": "FEF3C7",
    "Admin":    "EDE9FE",
    "User":     "DBEAFE",
}

thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, headers, col_widths):
    """Write a styled header row."""
    for col_i, (text, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col_i, value=text)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[row].height = 28


def data_cell(ws, row, col, value, bg=None, wrap=False, align="left"):
    cell = ws.cell(row=row, column=col, value=value if value is not None else "—")
    cell.font = Font(size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def sheet_title(ws, title, subtitle=""):
    ws.merge_cells("A1:Z1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=14, color=DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells("A2:Z2")
        c2 = ws["A2"]
        c2.value = subtitle
        c2.font = Font(italic=True, size=9, color="6B7280")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16
        return 3
    return 2


def freeze(ws, row, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


# ─────────────────────────────────
# WORKBOOK
# ─────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default blank sheet

generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

# ══════════════════════════════════
# SHEET 1 — OVERVIEW
# ══════════════════════════════════
ws = wb.create_sheet("📊 Overview")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 42

ws["A1"].value = "AryVerse — Creators System Export"
ws["A1"].font = Font(bold=True, size=16, color=DARK)
ws.row_dimensions[1].height = 34

ws["A2"].value = f"Generated: {generated_at}   |   Supabase DB snapshot"
ws["A2"].font = Font(italic=True, size=9, color="6B7280")
ws.row_dimensions[2].height = 18

# Summary table headers
for r, (label, val) in enumerate([
    ("Table", "Rows"),
    ("Profiles (Users)",        str(len(PROFILES))),
    ("Tasks",                   str(len(TASKS))),
    ("Points Log",              str(len(POINTS_LOG))),
    ("Activity Log",            str(len(ACTIVITY_LOG))),
    ("Department Access Rules", str(len(DEPARTMENT_ACCESS))),
    ("Password Reset Requests", str(len(PASSWORD_RESET_REQUESTS))),
], start=4):
    c1 = ws.cell(row=r, column=1, value=label)
    c2 = ws.cell(row=r, column=2, value=val)
    if r == 4:
        for c in [c1, c2]:
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor=PURPLE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        ws.row_dimensions[r].height = 24
    else:
        bg = PURPLE_LT if r % 2 == 1 else WHITE
        c1.fill = c2.fill = PatternFill("solid", fgColor=bg)
        c1.font = c2.font = Font(size=10)
        c1.border = c2.border = BORDER
        c2.alignment = Alignment(horizontal="center")

# Stats block
stats_row = 12
ws.cell(row=stats_row, column=1).value = "Quick Stats"
ws.cell(row=stats_row, column=1).font = Font(bold=True, size=11, color=DARK)

stats = [
    ("Total Team Members",    len(PROFILES)),
    ("Directors",             sum(1 for p in PROFILES if p["role"]=="Director")),
    ("Admins",                sum(1 for p in PROFILES if p["role"]=="Admin")),
    ("Users",                 sum(1 for p in PROFILES if p["role"]=="User")),
    ("Active Tasks",          sum(1 for t in TASKS if t["status"]=="Pending")),
    ("Completed Tasks",       sum(1 for t in TASKS if t["status"]=="Completed")),
    ("Total Tokens Awarded",  sum(p["total_tokens"] for p in PROFILES)),
    ("Total Token Events",    len(POINTS_LOG)),
    ("Departments",           len(set(p["department"] for p in PROFILES if p["department"]))),
]
for i, (lbl, val) in enumerate(stats):
    r = stats_row + 1 + i
    c1 = ws.cell(row=r, column=1, value=lbl)
    c2 = ws.cell(row=r, column=2, value=val)
    bg = PURPLE_LT if i % 2 == 0 else WHITE
    c1.fill = c2.fill = PatternFill("solid", fgColor=bg)
    c1.font = Font(size=10)
    c2.font = Font(bold=True, size=10)
    c1.border = c2.border = BORDER
    c2.alignment = Alignment(horizontal="center")


# ══════════════════════════════════
# SHEET 2 — PROFILES
# ══════════════════════════════════
ws = wb.create_sheet("👥 Team Profiles")
ws.sheet_view.showGridLines = False
start = sheet_title(ws, "Team Profiles", f"All employees — {len(PROFILES)} records")

headers = ["Emp ID","Full Name","Role","Department","Email","Phone","Date of Birth",
           "Total Tokens","Banked Min","LinkedIn","GitHub","Resume Link","Created At"]
widths  = [14, 26, 12, 22, 30, 16, 14, 13, 12, 14, 14, 14, 18]
hdr(ws, start, headers, widths)
freeze(ws, start+1)

for i, p in enumerate(PROFILES):
    r = start + 1 + i
    bg_row = ROLE_COLORS.get(p["role"], WHITE)
    cols = [
        p["employee_id"], p["full_name"], p["role"], p["department"],
        p["email"], p["phone"], p["date_of_birth"],
        p["total_tokens"], p["banked_minutes"],
        p["linkedin_url"], p["github_url"], p["resume_url"],
        p["created_at"],
    ]
    for col_i, val in enumerate(cols, start=1):
        bg = bg_row if col_i <= 4 else (PURPLE_LT if i % 2 == 0 else WHITE)
        data_cell(ws, r, col_i, val, bg=bg, align="center" if col_i in [1,3,7,8,9,13] else "left")
        if col_i in [10,11,12] and val and val != "—":
            cell = ws.cell(row=r, column=col_i)
            cell.value = "🔗 Link"
            cell.hyperlink = str(val)
            cell.font = Font(color="1D4ED8", underline="single", size=9)


# ══════════════════════════════════
# SHEET 3 — LEADERBOARD
# ══════════════════════════════════
ws = wb.create_sheet("🏆 Leaderboard")
ws.sheet_view.showGridLines = False
ranked = sorted([p for p in PROFILES if p["role"]=="User"], key=lambda x: x["total_tokens"], reverse=True)
start = sheet_title(ws, "Token Leaderboard", f"Users ranked by total tokens — {len(ranked)} users")

headers = ["Rank","Emp ID","Full Name","Department","Total Tokens","Token Credit Balance","Banked Minutes","Joined"]
widths  = [7, 14, 26, 22, 14, 20, 14, 18]
hdr(ws, start, headers, widths)
freeze(ws, start+1)

medal = {1:"🥇", 2:"🥈", 3:"🥉"}
for i, p in enumerate(ranked):
    r = start + 1 + i
    rank = i + 1
    bg = ("FEF9C3" if rank==1 else "F1F5F9" if rank==2 else "FEF3C7" if rank==3 else
          PURPLE_LT if i % 2 == 0 else WHITE)
    data_cell(ws, r, 1, f"{medal.get(rank,rank)} {rank}", bg=bg, align="center")
    data_cell(ws, r, 2, p["employee_id"], bg=bg, align="center")
    data_cell(ws, r, 3, p["full_name"], bg=bg)
    data_cell(ws, r, 4, p["department"], bg=bg)
    data_cell(ws, r, 5, p["total_tokens"], bg=bg, align="center")
    ws.cell(row=r, column=5).font = Font(bold=True, size=10, color=("7C3AED" if p["total_tokens"]>0 else "6B7280"))
    data_cell(ws, r, 6, round(p["token_credit_balance"],2), bg=bg, align="center")
    data_cell(ws, r, 7, p["banked_minutes"], bg=bg, align="center")
    data_cell(ws, r, 8, p["created_at"], bg=bg, align="center")


# ══════════════════════════════════
# SHEET 4 — TASKS
# ══════════════════════════════════
ws = wb.create_sheet("📋 Tasks")
ws.sheet_view.showGridLines = False
start = sheet_title(ws, "All Tasks", f"{len(TASKS)} tasks — sorted by creation date (newest first)")

headers = ["Title","Assigned To","Assignee ID","Created By","Status","Tokens","Dir. Approved",
           "Deadline","Submitted At","Approved At","GitHub Issue","Issue State",
           "Submission Note","Admin Feedback","Created At"]
widths  = [36, 26, 12, 22, 14, 8, 14, 18, 18, 18, 30, 13, 50, 50, 18]
hdr(ws, start, headers, widths)
freeze(ws, start+1)

for i, t in enumerate(TASKS):
    r = start + 1 + i
    status_bg = STATUS_COLORS.get(t["status"], WHITE)
    row_bg = PURPLE_LT if i % 2 == 0 else WHITE
    cols = [
        t["title"], t["assigned_to"], t["assignee_id"], t["created_by"],
        t["status"], t["tokens"],
        "✅ Yes" if t["director_approved"] else "⏳ No",
        t["deadline"], t["submitted_at"], t["approved_at"],
        t.get("pow_url"), t.get("issue_state"),
        t["submission_note"], t["admin_feedback"], t["created_at"],
    ]
    for col_i, val in enumerate(cols, start=1):
        if col_i == 5:
            data_cell(ws, r, col_i, val, bg=status_bg, align="center")
        elif col_i == 11 and val:
            cell = ws.cell(row=r, column=col_i)
            cell.value = "🔗 View Issue"
            cell.hyperlink = str(val)
            cell.font = Font(color="1D4ED8", underline="single", size=9)
            cell.border = BORDER
            cell.fill = PatternFill("solid", fgColor=row_bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_i in [13, 14]:
            data_cell(ws, r, col_i, val, bg=row_bg, wrap=True, align="left")
        elif col_i in [6, 7, 8, 9, 10, 12, 15]:
            data_cell(ws, r, col_i, val, bg=row_bg, align="center")
        else:
            data_cell(ws, r, col_i, val, bg=row_bg)
    ws.row_dimensions[r].height = 40


# ══════════════════════════════════
# SHEET 5 — POINTS LOG
# ══════════════════════════════════
ws = wb.create_sheet("💰 Points Log")
ws.sheet_view.showGridLines = False
start = sheet_title(ws, "Points / Token Log", f"{len(POINTS_LOG)} token events")

headers = ["Emp ID","Name","Task Title","Tokens Awarded","Reason","Awarded At"]
widths  = [14, 26, 40, 15, 40, 20]
hdr(ws, start, headers, widths)
freeze(ws, start+1)

for i, p in enumerate(POINTS_LOG):
    r = start + 1 + i
    bg = GREEN_BG if i % 2 == 0 else WHITE
    data_cell(ws, r, 1, p["employee_id"], bg=bg, align="center")
    data_cell(ws, r, 2, p["full_name"], bg=bg)
    data_cell(ws, r, 3, p["task_title"], bg=bg)
    c = data_cell(ws, r, 4, p["tokens_awarded"], bg=bg, align="center")
    c.font = Font(bold=True, size=10, color="065F46")
    data_cell(ws, r, 5, p["reason"], bg=bg)
    data_cell(ws, r, 6, p["created_at"], bg=bg, align="center")


# ══════════════════════════════════
# SHEET 6 — ACTIVITY LOG
# ══════════════════════════════════
ws = wb.create_sheet("📜 Activity Log")
ws.sheet_view.showGridLines = False
start = sheet_title(ws, "Activity Log", f"{len(ACTIVITY_LOG)} events — ordered newest first")

ACTION_COLORS = {
    "task_approved":       "D1FAE5",
    "task_rejected":       "FEE2E2",
    "task_reassigned":     "FEF9C3",
    "task_marked_done":    "DBEAFE",
    "task_assigned":       "EDE9FE",
    "director_approved_task":"E0F2FE",
    "task_deleted":        "FFE4E6",
    "user_added":          "ECFDF5",
    "custom_message":      "FFF7ED",
    "tokens_given":        "FEF9C3",
}

headers = ["Timestamp","Actor","Actor Role","Target User","Task","Action Type","Message"]
widths  = [18, 26, 12, 26, 36, 24, 70]
hdr(ws, start, headers, widths)
freeze(ws, start+1)

for i, a in enumerate(ACTIVITY_LOG):
    r = start + 1 + i
    bg = ACTION_COLORS.get(a["action_type"], WHITE)
    data_cell(ws, r, 1, a["created_at"], bg=bg, align="center")
    data_cell(ws, r, 2, a["actor_name"], bg=bg)
    data_cell(ws, r, 3, a["actor_role"], bg=bg, align="center")
    data_cell(ws, r, 4, a["target_user"], bg=bg)
    data_cell(ws, r, 5, a["task_title"], bg=bg)
    data_cell(ws, r, 6, a["action_type"], bg=bg, align="center")
    data_cell(ws, r, 7, a["message"], bg=bg, wrap=True)
    ws.row_dimensions[r].height = 32


# ══════════════════════════════════
# SHEET 7 — DEPARTMENT ACCESS
# ══════════════════════════════════
ws = wb.create_sheet("🏢 Dept Access")
ws.sheet_view.showGridLines = False
start = sheet_title(ws, "Department Access Rules", f"{len(DEPARTMENT_ACCESS)} cross-department viewing grants")

headers = ["User Emp ID","User Name","Their Dept","Can View Dept","Granted By","Granted At"]
widths  = [14, 26, 22, 22, 26, 20]
hdr(ws, start, headers, widths)
freeze(ws, start+1)

for i, da in enumerate(DEPARTMENT_ACCESS):
    r = start + 1 + i
    bg = BLUE_BG if i % 2 == 0 else WHITE
    data_cell(ws, r, 1, da["user_emp_id"], bg=bg, align="center")
    data_cell(ws, r, 2, da["user_name"], bg=bg)
    data_cell(ws, r, 3, da["department"], bg=bg, align="center")
    data_cell(ws, r, 4, da["can_view_department"], bg=bg, align="center")
    data_cell(ws, r, 5, da["granted_by_name"], bg=bg)
    data_cell(ws, r, 6, da["created_at"], bg=bg, align="center")


# ══════════════════════════════════
# SHEET 8 — PASSWORD RESET REQUESTS
# ══════════════════════════════════
ws = wb.create_sheet("🔑 Password Resets")
ws.sheet_view.showGridLines = False
start = sheet_title(ws, "Password Reset Requests", f"{len(PASSWORD_RESET_REQUESTS)} requests")

headers = ["Email","Status","Resolved By","Requested At","Resolved At"]
widths  = [34, 14, 26, 22, 22]
hdr(ws, start, headers, widths)
freeze(ws, start+1)

for i, pr in enumerate(PASSWORD_RESET_REQUESTS):
    r = start + 1 + i
    bg = STATUS_COLORS.get(pr["status"], WHITE)
    data_cell(ws, r, 1, pr["email"], bg=bg)
    data_cell(ws, r, 2, pr["status"].capitalize(), bg=bg, align="center")
    data_cell(ws, r, 3, pr["resolved_by_name"], bg=bg)
    data_cell(ws, r, 4, pr["created_at"], bg=bg, align="center")
    data_cell(ws, r, 5, pr["resolved_at"], bg=bg, align="center")


# ─────────────────────────────────
# SAVE
# ─────────────────────────────────
out_path = os.path.join(os.path.dirname(__file__), "..", "CreatorsSystem_Export.xlsx")
out_path = os.path.normpath(out_path)
wb.save(out_path)
print(f"\n✅  Saved → {out_path}\n")
print("Sheets exported:")
for sh in wb.sheetnames:
    print(f"  • {sh}")
